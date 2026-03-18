from __future__ import annotations

from io import BytesIO
from re import sub

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from apps.core.attendance.services import table_pdf_bytes


def _filename(value):
    return sub(r'[^a-zA-Z0-9_-]+', '-', value.strip().lower()).strip('-') or 'report'


def workbook_bytes(report):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Report'

    row_index = 1
    sheet.cell(row=row_index, column=1, value=report['title'])
    sheet.cell(row=row_index, column=1).font = Font(bold=True, size=14)
    row_index += 2

    if report.get('summary'):
        for label, value in report['summary']:
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=str(value))
            row_index += 1
        row_index += 1

    for column_index, header in enumerate(report['columns'], start=1):
        sheet.cell(row=row_index, column=column_index, value=header)
        sheet.cell(row=row_index, column=column_index).font = Font(bold=True)
    row_index += 1

    for row in report['rows']:
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=str(value))
        row_index += 1

    for column_cells in sheet.columns:
        width = 18
        for cell in column_cells:
            width = max(width, len(str(cell.value or '')) + 2)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 40)

    sheet.freeze_panes = 'A4'
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def pdf_bytes(report):
    rows = list(report['rows'])
    if report.get('summary'):
        rows = [[label, value] for label, value in report['summary']] + [['', '']] + rows
    return table_pdf_bytes(report['title'], report['columns'], rows)


def export_response(report, export_format):
    base_name = _filename(report['title'])
    if export_format == 'xlsx':
        content = workbook_bytes(report)
        response = HttpResponse(
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{base_name}.xlsx"'
        return response

    if export_format == 'pdf':
        content = pdf_bytes(report)
        response = HttpResponse(content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{base_name}.pdf"'
        return response

    return None
